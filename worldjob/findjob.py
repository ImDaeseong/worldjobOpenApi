from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook


def loadurl():
    spage = 'http://www.worldjob.or.kr/openapi/openapi.do?'  # https://www.data.go.kr/dataset/3038249/openapi.do
    sdobType = '1'  # 1:해외취업,2:해외연수,3:해외인턴,4:해외봉사,5:해외창업
    sdsptcKsco = '01'  # 직종별코드(해외취업,연수만 사용)01:전산,컴퓨터,02:전기/전자,06:기계/금속,07:건설/토목,08:사무/서비스,09:의료,10:기타
    scontinent = '1'  # 대륙별코드 1:아시아,2:북아메리카, 3:남아메리카,4:유럽,5:오세아니아,6:아프리카
    sepmt61 = 'Y'  # 일자리Best20(해외취업만 사용)Y,N
    sshowItemListCount = '1000'  # 한번에보여질리스트갯수출력결과
    url = "{0}dobType={1}&dsptcKsco={2}&continent={3}&showItemListCount={4}&sepmt61={5}".format(spage, sdobType,
                                                                                                sdsptcKsco, scontinent,
                                                                                                sshowItemListCount,
                                                                                                sepmt61)
    # print(url)
    responses = requests.get(url)
    soup = BeautifulSoup(responses.content, 'lxml-xml')

    for worldjob in soup.findAll('WORLDJOB'):
        ERR_CD = worldjob.find('ERR_CD')
        COUNT = worldjob.find('COUNT')
        print("ERR_CD: " + ERR_CD.string + " COUNT: " + COUNT.string)

        for item in worldjob.find_all('ITEM'):
            rctntcSj = item.find('rctntcSj')
            rctntcSprtQualfCn = item.find('rctntcSprtQualfCn')
            dsptcNationScd = item.find('dsptcNationScd')
            dsptcKsco = item.find('dsptcKsco')
            joDemandCareerStleScd = item.find('joDemandCareerStleScd')
            joDemandAcdmcrScd = item.find('joDemandAcdmcrScd')
            rctntcEndDay = item.find('rctntcEndDay')
            linkUrl = item.find('linkUrl')
            directApply = item.find('directApply')

            if rctntcSj.string.find('일본') > 0:
                print("rctntcSj: " + rctntcSj.string)
                print("linkUrl: " + linkUrl.string)


def loadurl_excel():
    spage = 'http://www.worldjob.or.kr/openapi/openapi.do?'  # https://www.data.go.kr/dataset/3038249/openapi.do
    sdobType = '1'  # 1:해외취업,2:해외연수,3:해외인턴,4:해외봉사,5:해외창업
    sdsptcKsco = '01'  # 직종별코드(해외취업,연수만 사용)01:전산,컴퓨터,02:전기/전자,06:기계/금속,07:건설/토목,08:사무/서비스,09:의료,10:기타
    scontinent = '1'  # 대륙별코드 1:아시아,2:북아메리카, 3:남아메리카,4:유럽,5:오세아니아,6:아프리카
    sepmt61 = 'Y'  # 일자리Best20(해외취업만 사용)Y,N
    sshowItemListCount = '1000'  # 한번에보여질리스트갯수출력결과
    url = "{0}dobType={1}&dsptcKsco={2}&continent={3}&showItemListCount={4}&sepmt61={5}".format(spage, sdobType,
                                                                                                sdsptcKsco, scontinent,
                                                                                                sshowItemListCount,
                                                                                                sepmt61)
    # print(url)
    responses = requests.get(url)
    soup = BeautifulSoup(responses.content, 'lxml-xml')

    for worldjob in soup.findAll('WORLDJOB'):
        ERR_CD = worldjob.find('ERR_CD')
        COUNT = worldjob.find('COUNT')
        print("ERR_CD: " + ERR_CD.string + " COUNT: " + COUNT.string)

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'world job'

        row = 1
        for item in worldjob.find_all('ITEM'):
            rctntcSj = item.find('rctntcSj')
            rctntcSprtQualfCn = item.find('rctntcSprtQualfCn')
            dsptcNationScd = item.find('dsptcNationScd')
            dsptcKsco = item.find('dsptcKsco')
            joDemandCareerStleScd = item.find('joDemandCareerStleScd')
            joDemandAcdmcrScd = item.find('joDemandAcdmcrScd')
            rctntcEndDay = item.find('rctntcEndDay')
            linkUrl = item.find('linkUrl')
            directApply = item.find('directApply')

            sheet.cell(row, column=1).value = rctntcSj.string
            sheet.cell(row, column=2).value = rctntcSprtQualfCn.string
            sheet.cell(row, column=3).value = dsptcNationScd.string
            sheet.cell(row, column=4).value = dsptcKsco.string
            sheet.cell(row, column=5).value = joDemandCareerStleScd.string
            sheet.cell(row, column=6).value = joDemandAcdmcrScd.string
            sheet.cell(row, column=7).value = rctntcEndDay.string
            sheet.cell(row, column=8).value = linkUrl.string
            sheet.cell(row, column=9).value = directApply.string

            row += 1

        wb.save('e:/a.xlsx')


if __name__ == "__main__":
    loadurl()
    # loadurl_excel()
